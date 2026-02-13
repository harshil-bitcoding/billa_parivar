import csv
import io
import os
import openpyxl
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from django.utils import timezone
from .models import Surname, Person, District, Taluka, Village, Country, DemoPerson, DemoParentChildRelation, ParentChildRelation

class LocationResolverService:
    @staticmethod
    def resolve_location(district_name, taluka_name, village_name):
        """
        Resolves hierarchical location (District -> Taluka -> Village).
        Strict policy: Returns None if match not found.
        Matching is case-insensitive.
        """
        # 1. Resolve District
        district_name = str(district_name).strip()
        district = District.objects.filter(name__iexact=district_name).first()
        if not district:
            return None, f"District Not Found: '{district_name}'"

        # 2. Resolve Taluka
        taluka_name = str(taluka_name).strip()
        taluka = Taluka.objects.filter(name__iexact=taluka_name, district=district).first()
        if not taluka:
            return None, f"Taluka Not Found: '{taluka_name}' (in District: {district_name})"

        # 3. Resolve Village
        village_name = str(village_name).strip()
        village = Village.objects.filter(name__iexact=village_name, taluka=taluka).first()
        if not village:
            return None, f"Village Not Found: '{village_name}' (in Taluka: {taluka_name})"
        
        return village, "exact"

class CSVImportService:
    @staticmethod
    def clean_val(val):
        if val is None:
            return ""
            
        # Handle direct float/int from Excel
        if isinstance(val, (int, float)):
            # Check if it's effectively an integer (e.g. 123.0)
            if val == int(val):
                return str(int(val))
        
        val = str(val).strip()
        # Remove Excel formula wrapper ="value"
        if val.startswith('="') and val.endswith('"'):
            return val[2:-1]
        # Remove single quote prefix
        if val.startswith("'"):
            return val[1:]
            
        # Handle string "123.0" -> "123" (Common CSV/Excel artifact)
        if val.endswith(".0"):
            try:
                return str(int(float(val)))
            except ValueError:
                pass
                
        return val

    @staticmethod
    def resolve_surname(name):
        """
        Policy:
        Compare sheet data and DB data in the same case.
        If a case-insensitive match exists, use it to avoid unnecessary bug files.
        """
        name = str(name).strip()
        if not name:
            return None, "empty"
        
        # Use Case-insensitive match (PostgreSQL iexact handles "same case" comparison)
        obj = Surname.objects.filter(name__iexact=name).first()
        if obj:
            # Match found (e.g. "Patel" for "patel") - use it.
            return obj, "exact"
        
        # No match at all - create new
        new_surname = Surname.objects.create(name=name)
        return new_surname, "created"

    @classmethod
    def process_file(cls, uploaded_file, request=None, is_demo=False):
        """
        Core logic for processing CSV/XLSX files.
        Mirrored from DemoCSVUploadAPIView with added surname policy.
        """
        # 1. Save original file
        fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'uploads', 'original'))
        filename = fs.save(uploaded_file.name, uploaded_file)
        
        # Rewind file for processing
        uploaded_file.seek(0)
        
        ext = uploaded_file.name.lower().split('.')[-1]
        all_sheets_data = {}

        # 2. Extract Data
        if ext in ['xlsx', 'xls']:
            try:
                wb = openpyxl.load_workbook(uploaded_file, data_only=True)
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        rows.append(list(row))
                    all_sheets_data[sheet_name] = rows
            except Exception as e:
                return {"error": f"Failed to read XLSX: {str(e)}"}
        else:
            try:
                file_data = uploaded_file.read()
                try:
                    decoded = file_data.decode('utf-8-sig')
                except:
                    decoded = file_data.decode('latin-1')
                normalized = "\n".join(decoded.splitlines())
                reader = csv.reader(io.StringIO(normalized))
                rows = [row for row in reader]
                all_sheets_data['default'] = rows
            except Exception as e:
                return {"error": f"Failed to read CSV: {str(e)}"}

        total_created = 0
        total_updated = 0
        bug_rows = []
        
        # 3. Process Sheets
        global_village = None
        global_d_name, global_t_name, global_v_name = "", "", ""

        for s_idx, (sheet_name, rows) in enumerate(all_sheets_data.items()):
            # 3. Tab Processing Policy
            if s_idx == 0: # 1st Tab: Dashboard
                header_row_idx = -1
                for i in range(min(5, len(rows))):
                    row_str = " ".join([str(c) for c in rows[i] if c]).lower()
                    if 'district' in row_str and 'village' in row_str:
                        header_row_idx = i
                        break
                
                if header_row_idx != -1 and len(rows) > header_row_idx + 1:
                    data_row = rows[header_row_idx + 1]
                    # Assuming standard layout for Dashboard: District (0), Taluka (1), Village (2), RefCode (3)
                    global_d_name = cls.clean_val(data_row[0]) if len(data_row) > 0 else ""
                    global_t_name = cls.clean_val(data_row[1]) if len(data_row) > 1 else ""
                    global_v_name = cls.clean_val(data_row[2]) if len(data_row) > 2 else ""
                    
                    global_village, loc_status = LocationResolverService.resolve_location(global_d_name, global_t_name, global_v_name)
                    if not global_village:
                        return {"error": f"Dashboard Location Error: {loc_status} for {global_d_name}/{global_t_name}/{global_v_name}"}
                    
                    # Update Referral Code
                    ref_code = cls.clean_val(data_row[3]) if len(data_row) > 3 else ""
                    if ref_code:
                        global_village.referral_code = ref_code
                        global_village.save()
                continue

            if s_idx == 1: # 2nd Tab: Dummy
                continue

            # 3rd Tab & Onwards: Person Data (Surname Tabs as Master)
            if not global_village:
                continue

            # Master Tab Rule: Sheet name must exactly match a Surname in DB
            sheet_surname_obj = Surname.objects.filter(name__iexact=sheet_name.strip()).first()
            if not sheet_surname_obj:
                # If the sheet name doesn't match a surname, we skip this sheet as per "Master Tab" policy
                # Optionally log this as a bug if it's unexpected
                bug_rows.append([f"Sheet: {sheet_name}", "Skipped: Tab name does not match any Surname in database"])
                continue

            col_map = {}
            keywords = {
                'first_name': ['Firstname (In English)', 'Firstname', 'First name', 'In English', 'In emglish', 'In rmglish'],
                'guj_first_name': ['Firstname (In Gujarati)', 'In Gujarati', 'In Gujaratio', 'In Gujaration', 'In Gujaral', 'In Gujaralt', 'In Gujarai'],
                'middle_name': ['Father name (In English)', 'Father name', 'Name of Father', 'In English', 'In emglish', 'In rmglish'],
                'guj_middle_name': ['Father name (In Gujarati)', 'Father name Gujarati', 'In Gujarati', 'In Gujaratio', 'In Gujaration', 'In Gujaral', 'In Gujaralt', 'In Gujarai'],
                'surname': ['Surname', 'Sirname'],
                'mobile1': ['Mobile Number Main', 'Main', 'Mobile'],
                'mobile2': ['Mobile Number (Optional)', 'Secondary', 'Optional'],
                'dob': ['Birth Date', 'Birt Date', 'DOB'],
                'country': ['Country Name', 'Outside India', 'Country'],
                'int_mobile': ['International Mobile', 'International'],
                'profile': ['Profile', 'Profile Pic', 'Image'],
                'thumb_profile': ['Thumb profile', 'Thumb', 'Thumbnail'],
            }

            # Smart Header Detection
            header_row_idx = -1
            for i in range(min(10, len(rows))):
                row = rows[i]
                row_str = " ".join([str(c).strip() if c else "" for c in row]).lower()
                if any(x in row_str for x in ['firstname', 'surname', 'mobile']):
                    header_row_idx = i
                    current_parent = ""
                    for j, cell in enumerate(row):
                        cell_val = str(cell).strip() if cell else ""
                        if cell_val:
                            current_parent = cell_val
                        
                        next_cell_val = ""
                        if i + 1 < len(rows) and j < len(rows[i+1]):
                            next_cell_val = str(rows[i+1][j]).strip() if rows[i+1][j] else ""
                        
                        combined_col_str = f"{current_parent} {next_cell_val}".strip().lower()

                        if any(k in combined_col_str for k in ['firstname', 'first name']):
                            if any(k in combined_col_str for k in ['gujarati', 'gujrai', 'gujara', 'gujal', 'gujalt']):
                                col_map['guj_first_name'] = j
                            elif 'english' in combined_col_str or 'first_name' not in col_map:
                                col_map['first_name'] = j
                        elif any(k in combined_col_str for k in ['father', 'middle', 'parent']):
                            if any(k in combined_col_str for k in ['gujarati', 'gujrai', 'gujara', 'gujal', 'gujalt']):
                                col_map['guj_middle_name'] = j
                            elif 'english' in combined_col_str or 'middle_name' not in col_map:
                                col_map['middle_name'] = j
                        
                        # 2. Other Fields
                        for key, keys in keywords.items():
                            if key in ['gender', 'mobile1', 'mobile2', 'dob', 'country', 'int_mobile', 'surname', 'profile', 'thumb_profile']:
                                if any(k.lower() in combined_col_str for k in keys):
                                    if key == 'mobile1':
                                        if "main" in combined_col_str: col_map['mobile1'] = j
                                        elif "optional" in combined_col_str: col_map['mobile2'] = j
                                        elif 'mobile1' not in col_map: col_map['mobile1'] = j
                                    else:
                                        if any(k.lower() == combined_col_str or k.lower() == next_cell_val.lower() for k in keys):
                                            col_map[key] = j
                                        elif key not in col_map:
                                            col_map[key] = j
                    break

            if header_row_idx == -1:
                continue

            start_row = header_row_idx + 1
            if start_row < len(rows):
                sub_row_str = " ".join([str(x) for x in rows[start_row]]).lower()
                if any(x in sub_row_str for x in ["english", "gujarati", "main", "optional"]):
                    start_row += 1

            for idx, row in enumerate(rows[start_row:]):
                if not any(row): continue
                
                try:
                    # Person Data Extraction
                    f_name = cls.clean_val(row[col_map['first_name']]) if 'first_name' in col_map else ""
                    m_name = cls.clean_val(row[col_map['middle_name']]) if 'middle_name' in col_map else ""
                    guj_f_name = cls.clean_val(row[col_map['guj_first_name']]) if 'guj_first_name' in col_map else ""
                    guj_m_name = cls.clean_val(row[col_map['guj_middle_name']]) if 'guj_middle_name' in col_map else ""
                    s_name = cls.clean_val(row[col_map['surname']]) if 'surname' in col_map else ""
                    mob1 = cls.clean_val(row[col_map['mobile1']]) if 'mobile1' in col_map else ""
                    mob2 = cls.clean_val(row[col_map['mobile2']]) if 'mobile2' in col_map else ""
                    dob_raw = cls.clean_val(row[col_map['dob']]) if 'dob' in col_map else ""
                    country_name = cls.clean_val(row[col_map['country']]) if 'country' in col_map else ""
                    int_mob = cls.clean_val(row[col_map['int_mobile']]) if 'int_mobile' in col_map else ""
                    profile_path = cls.clean_val(row[col_map['profile']]) if 'profile' in col_map else ""
                    thumb_profile_path = cls.clean_val(row[col_map['thumb_profile']]) if 'thumb_profile' in col_map else ""
                    
                    # 1. Empty Row Check: Skip silently if the row has absolutely no data
                    # This avoids false "Surname Mismatch" bugs for blank rows.
                    if not any([f_name, m_name, guj_f_name, guj_m_name, s_name, mob1, mob2, dob_raw, country_name, int_mob]):
                        continue

                    # 2. Strict Surname Mismatch: Check if row data belongs in this Master Tab
                    # Rows with data but a different surname are logged to the bug file.
                    if s_name.strip().lower() != sheet_name.strip().lower():
                        bug_rows.append(list(row) + [f"Surname mismatch: Sheet is '{sheet_name}', but row says '{s_name}'"])
                        continue

                    # 3. Proceed with record creation/update (Partial data is allowed if identifying info exists)
                    surname_obj = sheet_surname_obj

                    # DOB Normalization
                    dob = str(dob_raw).strip() if dob_raw else ""
                    if dob:
                        # Handle YYYY-MM-DD HH:MM:SS (common in Excel/openpyxl)
                        if ' ' in dob:
                            dob = dob.split(' ')[0]
                        
                        if '-' in dob:
                            parts = dob.split('-')
                            if len(parts) == 3 and len(parts[0]) == 4: # YYYY-MM-DD
                                dob = f"{parts[2]}-{parts[1]}-{parts[0]}"

                    # Country Handling
                    c_name = country_name if country_name else "India"
                    country_obj, _ = Country.objects.get_or_create(name=c_name)
                    is_out = country_obj.name.lower() != 'india'

                    person_model = DemoPerson if is_demo else Person

                    person_defaults = {
                        'first_name': f_name,
                        'middle_name': m_name,
                        'guj_first_name': guj_f_name,
                        'guj_middle_name': guj_m_name,
                        'surname': surname_obj,
                        'date_of_birth': dob,
                        'mobile_number2': mob2,
                        'is_out_of_country': is_out,
                        'out_of_country': country_obj,
                        'international_mobile_number': int_mob,
                        'village': global_village,
                        'taluka': global_village.taluka if global_village else None,
                        'district': global_village.taluka.district if global_village and global_village.taluka else None,
                        'flag_show': True,
                        'is_demo': is_demo
                    }

                    if profile_path:
                        # Clean path if starts with /media/
                        if profile_path.startswith('/media/'):
                            profile_path = profile_path.replace('/media/', '', 1)
                        if profile_path.split('://')[0] not in ['http', 'https']:
                            person_defaults['profile'] = profile_path

                    if thumb_profile_path:
                        # Clean path if starts with /media/
                        if thumb_profile_path.startswith('/media/'):
                            thumb_profile_path = thumb_profile_path.replace('/media/', '', 1)
                        if thumb_profile_path.split('://')[0] not in ['http', 'https']:
                            person_defaults['thumb_profile'] = thumb_profile_path
                    
                    if mob1:
                        # Update or create if mobile is present
                        person, created = person_model.objects.update_or_create(
                            mobile_number1=mob1, is_deleted=False, defaults=person_defaults
                        )
                    else:
                        # Success case: Mobile missing but name or other identifying info present
                        person_defaults['mobile_number1'] = mob1
                        person = person_model.objects.create(**person_defaults)
                        created = True

                    if created: total_created += 1
                    else: total_updated += 1
                except Exception as e:
                    bug_rows.append(list(row) + [str(e)])

        # 4. Process Relations (Name-based) - Mirrored from Demo
        person_model = DemoPerson if is_demo else Person
        rel_model = DemoParentChildRelation if is_demo else ParentChildRelation

        system_admin = person_model.objects.filter(is_super_admin=True, is_deleted=False).first()
        if not system_admin:
            system_admin = person_model.objects.filter(is_admin=True, is_deleted=False).first()
        
        if system_admin:
            # Mirror the Demo logic: check all persons for potential father matches
            all_persons = person_model.objects.filter(is_deleted=False).select_related('surname')
            for child in all_persons:
                father_name = child.middle_name
                if father_name and child.surname:
                    father = person_model.objects.filter(
                        first_name__iexact=father_name,
                        surname=child.surname,
                        is_deleted=False
                    ).exclude(id=child.id).first()
                    if father:
                        rel_model.objects.get_or_create(
                            parent=father, 
                            child=child,
                            defaults={'created_user': system_admin}
                        )

        # 5. Generate Bug CSV if needed
        bug_url = None
        if bug_rows:
            bug_fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'uploads', 'bugs'))
            bug_filename = f"bug_{timezone.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["Row Data", "Error Message"])
            for bug in bug_rows:
                writer.writerow(bug)
            
            bug_fs.save(bug_filename, io.BytesIO(output.getvalue().encode('utf-8')))
            
            if request:
                bug_url = request.build_absolute_uri(settings.MEDIA_URL + f"uploads/bugs/{bug_filename}")
            else:
                bug_url = settings.MEDIA_URL + f"uploads/bugs/{bug_filename}"

        return {
            "created": total_created,
            "updated": total_updated,
            "bug_file_url": bug_url,
            "bug_count": len(bug_rows),
            "original_filename": filename
        }
